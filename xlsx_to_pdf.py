#!/usr/bin/env python3
"""
Excel to PDF Converter

This script reads Excel files (.xlsx), reads specific sheets, and saves them to PDF format.
Supports multiple output formats and customization options.
"""

import os
import sys
import argparse
from pathlib import Path
import platform
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelToPdfConverter:
    """Convert Excel files to PDF format."""
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls']
        self.default_font = self._register_preferred_windows_font()
        self.page_size = A4

    def _register_preferred_windows_font(self):
        """Register a Unicode font on Windows to avoid missing diacritics and return its name.

        Falls back to Helvetica if no TTF font found.
        """
        fonts_tried = [
            ("Arial", "C:/Windows/Fonts/arial.ttf"),
            ("Calibri", "C:/Windows/Fonts/calibri.ttf"),
            ("Times New Roman", "C:/Windows/Fonts/times.ttf"),
            ("Segoe UI", "C:/Windows/Fonts/segoeui.ttf"),
        ]
        for display_name, path in fonts_tried:
            try:
                if os.path.exists(path):
                    pdfmetrics.registerFont(TTFont(display_name, path))
                    logger.info(f"Registered font for PDF: {display_name} -> {path}")
                    return display_name
            except Exception as register_error:
                logger.debug(f"Could not register font {display_name}: {register_error}")
        return "Helvetica"
    
    def list_sheets(self, excel_file):
        """List all sheets in the Excel file."""
        try:
            workbook = load_workbook(excel_file, read_only=True)
            sheets = workbook.sheetnames
            workbook.close()
            return sheets
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            return []
    
    def read_sheet(self, excel_file, sheet_name=None):
        """Read a specific sheet from Excel file."""
        try:
            if sheet_name:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
            else:
                # Read the first sheet if no sheet name specified
                df = pd.read_excel(excel_file)
                sheet_name = pd.ExcelFile(excel_file).sheet_names[0]
            
            logger.info(f"Successfully read sheet '{sheet_name}' with {len(df)} rows and {len(df.columns)} columns")
            return df, sheet_name
        except Exception as e:
            logger.error(f"Error reading sheet '{sheet_name}': {e}")
            return None, None
    
    def create_pdf_from_dataframe(self, df, output_file, title=None, font_name=None):
        """Create PDF from pandas DataFrame."""
        try:
            chosen_font = font_name or self.default_font
            doc = SimpleDocTemplate(output_file, pagesize=self.page_size)
            story = []
            
            # Add title
            if title:
                styles = getSampleStyleSheet()
                # Ensure title uses a Unicode-capable font
                styles['Title'].fontName = chosen_font
                title_paragraph = Paragraph(title, styles['Title'])
                story.append(title_paragraph)
                story.append(Spacer(1, 12))
            
            # Convert DataFrame to list of lists for table
            data = [df.columns.tolist()] + df.values.tolist()
            
            # Create table
            # Estimate column widths based on text length to reduce wrapping/clipping
            col_text_lengths = []
            for col_idx, col_name in enumerate(df.columns.tolist()):
                max_len = len(str(col_name))
                for value in df.iloc[:, col_idx].tolist():
                    max_len = max(max_len, len(str(value)))
                col_text_lengths.append(max_len)

            total_len = sum(col_text_lengths) if col_text_lengths else 1
            available_width = self.page_size[0] - 40  # margins approx
            col_widths = [max(40, available_width * (length / total_len)) for length in col_text_lengths]

            table = Table(data, colWidths=col_widths)
            
            # Style the table
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), chosen_font),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), chosen_font),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
            table.setStyle(style)
            
            story.append(table)
            doc.build(story)
            
            logger.info(f"PDF created successfully: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating PDF: {e}")
            return False
    
    def _convert_with_excel_com(self, input_file, output_file, sheet_name=None, visible=False):
        """Use Microsoft Excel COM automation to export to PDF (preserves formatting)."""
        try:
            import win32com.client  # type: ignore
        except Exception as exc:
            logger.error("win32com is required for Excel engine but not available: %s", exc)
            return False

        excel = None
        workbook = None
        try:
            excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = bool(visible)
            workbook = excel.Workbooks.Open(str(Path(input_file).resolve()))

            if sheet_name:
                try:
                    ws = workbook.Worksheets(sheet_name)
                except Exception:
                    logger.error(f"Sheet not found in Excel: {sheet_name}")
                    return False
                ws.ExportAsFixedFormat(0, str(Path(output_file).resolve()))
            else:
                # Export the first sheet only by default to match pandas behavior
                ws = workbook.Worksheets(1)
                ws.ExportAsFixedFormat(0, str(Path(output_file).resolve()))

            logger.info(f"Excel COM exported PDF successfully: {output_file}")
            return True
        except Exception as exc:
            logger.error(f"Excel COM export failed: {exc}")
            return False
        finally:
            try:
                if workbook is not None:
                    workbook.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass

    def convert_excel_to_pdf(self, input_file, output_file=None, sheet_name=None, engine: str = "auto", font_name: str | None = None, excel_visible: bool = False):
        """Main conversion function."""
        try:
            # Validate input file
            if not os.path.exists(input_file):
                logger.error(f"Input file not found: {input_file}")
                return False
            
            file_ext = Path(input_file).suffix.lower()
            if file_ext not in self.supported_formats:
                logger.error(f"Unsupported file format: {file_ext}")
                return False
            
            # Generate output filename if not provided
            if output_file is None:
                base_name = Path(input_file).stem
                sheet_suffix = f"_{sheet_name}" if sheet_name else ""
                output_file = f"{base_name}{sheet_suffix}.pdf"
            
            # Decide engine
            chosen_engine = engine
            if engine == "auto":
                if platform.system().lower().startswith("win"):
                    try:
                        import win32com.client  # noqa: F401
                        chosen_engine = "excel"
                    except Exception:
                        chosen_engine = "reportlab"
                else:
                    chosen_engine = "reportlab"

            if chosen_engine == "excel":
                success = self._convert_with_excel_com(input_file, output_file, sheet_name=sheet_name, visible=excel_visible)
                return success
            
            # Fallback/reportlab path
            df, actual_sheet_name = self.read_sheet(input_file, sheet_name)
            if df is None:
                return False
            
            title = f"Sheet: {actual_sheet_name}" if actual_sheet_name else "Excel Data"
            success = self.create_pdf_from_dataframe(df, output_file, title, font_name=font_name or self.default_font)
            
            return success
            
        except Exception as e:
            logger.error(f"Conversion failed: {e}")
            return False


def main():
    """Main function to handle command line arguments."""
    parser = argparse.ArgumentParser(description='Convert Excel files to PDF')
    parser.add_argument('input_file', help='Input Excel file path')
    parser.add_argument('-o', '--output', help='Output PDF file path')
    parser.add_argument('-s', '--sheet', help='Sheet name to convert (default: first sheet)')
    parser.add_argument('-l', '--list-sheets', action='store_true', help='List all sheets in the Excel file')
    parser.add_argument('--engine', choices=['auto', 'excel', 'reportlab'], default='auto', help='Conversion engine. excel preserves formatting (Windows only). reportlab is cross-platform. auto picks the best.')
    parser.add_argument('--font', help='Font name to use for PDF (reportlab engine). Defaults to a registered Windows font if available.')
    parser.add_argument('--excel-visible', action='store_true', help='Show Excel window during export (excel engine)')
    
    args = parser.parse_args()
    
    converter = ExcelToPdfConverter()
    
    # List sheets if requested
    if args.list_sheets:
        sheets = converter.list_sheets(args.input_file)
        if sheets:
            print(f"Available sheets in '{args.input_file}':")
            for i, sheet in enumerate(sheets, 1):
                print(f"  {i}. {sheet}")
        else:
            print("No sheets found or error reading file.")
        return
    
    # Convert Excel to PDF
    success = converter.convert_excel_to_pdf(
        input_file=args.input_file,
        output_file=args.output,
        sheet_name=args.sheet,
        engine=args.engine,
        font_name=args.font,
        excel_visible=args.excel_visible
    )
    
    if success:
        print(f"Conversion completed successfully!")
    else:
        print("Conversion failed. Check the logs for details.")
        sys.exit(1)


if __name__ == "__main__":
    main() 